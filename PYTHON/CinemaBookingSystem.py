from tabulate import tabulate
from openpyxl import Workbook, load_workbook

import random

# Defining movies and their prices
movies = {}
movies["1"] = {"title": "Ant-Man And The Wasp: Quantumania", "Genre": "Action/Adventure", "price": 680}
movies["2"] = {"title": "Avatar: The Way Of Water", "Genre": "Action/Adventure", "price": 680}
movies["3"] = {"title": "Sound Of Silence", "Genre": "Action/Adventure", "price": 540}
movies["4"] = {"title": "The First Slam Dunk", "Genre": "Anime", "price": 480}
movies["5"] = {"title": "Oras De Peligro", "Genre": "Drama", "price": 380}

# Defining the seats avalability 
seats = {
    movie_num: {
        "A": [i for i in range(1, 6)],
        "B": [i for i in range(1, 6)],
        "C": [i for i in range(1, 6)],
        "D": [i for i in range(1, 6)],
        "E": [i for i in range(1, 6)],
    } for movie_num in movies.keys()
    }

# Display the menu
def main():
    while True:     # Menu of the program and prompt the user to enter a choice
        print("\nWelcome to Cinema Booking System!")
        print("1. View Showings")
        print("2. Reserve Ticket")
        print("3. Exit Menu")
        choice = input("Enter choice (1-3): ")
        if choice == "1":       # Calls the function depending on the user's choice
            view_movies()
        elif choice == "2":
            reserve_ticket()
        elif choice == "3":
            print("\nThank you for using Cinema Booking System!")
            break
        else:
            print("Invalid choice, please try again.")

# Display avaliable seats for a particular movie
def print_seats(movie):
    print(f"\nAvailable seats for {movies[movie]['title']}:")
    table_data = [] 
    for row, nums in seats[movie].items():      # Loop through each seat number and store them in the table_data
        table_data.append([row] + [f"{num}" if num in nums else "   " for num in range(1, 6)])
    headers = [""] + [f"{i}" for i in range(1, 6)]
    print(tabulate(table_data, headers, tablefmt="fancy_grid")) 

# Display available movies
def view_movies():
    print("\nAvailable movies:")
    headers = ["  ", "Title", "Genre", "Price"]
    rows = []
    for key in movies.keys():
        title = movies[key]["title"]
        genre = movies[key]["Genre"]
        price = movies[key]["price"]        
        rows.append([key, title, genre, price])
    print(tabulate(rows, headers=headers, tablefmt="fancy_grid"))

# Function to reserve a ticket
def reserve_ticket():
    print(" ")      # Display list of movies and prompt user to select one
    print("\nPlease select a movie to book tickets for:")
    view_movies()
    movie = input("Enter movie number: ")
    while movie not in movies:
        print("Invalid movie number, please try again.")
        movie = input("Enter movie number: ")
    print(f"Selected movie: {movies[movie]['title']}")
    
    print_seats(movie)      # Ask user to select a seat and confirm selection
    seat = input("Enter row (A-E): ").upper()
    while seat not in seats[movie]:
        print("Invalid row, please try again.")
        seat = input("Enter row (A-E): ").upper()
    num = int(input("Enter seat number (1-5): "))
    while num not in seats[movie][seat]:
        print("Seat is not available, please choose another seat.")
        num = int(input("Enter seat number (1-5): "))
    print(" ")
    print(f"Selected seat: {seat}{num}")
    
    print(" ")      # Confirm the booking and display booking details
    confirm = input("Confirm booking? (Y/N): ").upper()
    while confirm not in ["Y", "N"]:
        print("Invalid option, please try again.")
        confirm = input("Confirm booking? (Y/N): ").upper()
    if confirm == "Y":
        seats[movie][seat][num-1] = " "
        generate_receipt(movie, seat, num)
        print(" ")
        print("Booking confirmed!")
    else:
        print("Booking canceled.")

# Generate receipt with random transaction number
def generate_receipt(movie, seat, num):
    transaction_number = random.randint(100000000, 999999999)
    print("\nYou made a reservation for: ")
    table = [["Transaction No:", transaction_number],
             ["Movie:", movies[movie]['title']],
             ["Seat:", seat+str(num)],
             ["Price:", f"₱{movies[movie]['price']}"]]
    print(tabulate(table, tablefmt="fancy_grid"))
    save_bookings(movie, seat, num, transaction_number)

# Save bookings in an Excel file
def save_bookings(movie, seat, num, transaction_number):
    try:
        wb = load_workbook("bookings.xlsx")
    except FileNotFoundError:
        wb = Workbook()
    sheet = wb.active
    if sheet.max_row == 0:
        sheet["A1"] = "Movie"
        sheet["B1"] = "Seat"
        sheet["C1"] = "Price"
        sheet["D1"] = "Transaction No."
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value=movies[movie]["title"])
    sheet.cell(row=row, column=2, value=f"{seat}{num}")
    sheet.cell(row=row, column=3, value=f"₱{movies[movie]['price']:.2f}")
    sheet.cell(row=row, column=4, value=transaction_number)
    wb.save("bookings.xlsx")

main()